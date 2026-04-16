import io
from datetime import date
from decimal import Decimal
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.enums import ReportFormat, TimeframeType, TransactionStatus
from app.models.match import Match
from app.models.transaction import Transaction

def resolve_periods(
    timeframe: TimeframeType,
    period: Optional[str] = None,
    quarter: Optional[int] = None,
    year: Optional[int] = None,
    from_period: Optional[str] = None,
    to_period: Optional[str] = None,
) -> list[str]:
    """Convert a timeframe selection into a sorted list of YYYY-MM period strings."""
    if timeframe == TimeframeType.single_month:
        if not period:
            raise ValueError("period is required for single_month timeframe")
        return [period]

    if timeframe == TimeframeType.quarter:
        if quarter is None or year is None:
            raise ValueError("quarter and year are required for quarter timeframe")
        if quarter not in (1, 2, 3, 4):
            raise ValueError("quarter must be 1–4")
        start_month = (quarter - 1) * 3 + 1
        return [f"{year}-{m:02d}" for m in range(start_month, start_month + 3)]

    if timeframe == TimeframeType.ytd:
        if year is None:
            year = date.today().year
        current_month = date.today().month if year == date.today().year else 12
        return [f"{year}-{m:02d}" for m in range(1, current_month + 1)]

    if timeframe == TimeframeType.full_year:
        if year is None:
            raise ValueError("year is required for full_year timeframe")
        return [f"{year}-{m:02d}" for m in range(1, 13)]

    if timeframe == TimeframeType.custom:
        if not from_period or not to_period:
            raise ValueError("from_period and to_period are required for custom timeframe")
        periods: list[str] = []
        fy, fm = int(from_period[:4]), int(from_period[5:7])
        ty, tm = int(to_period[:4]), int(to_period[5:7])
        cy, cm = fy, fm
        while (cy, cm) <= (ty, tm):
            periods.append(f"{cy}-{cm:02d}")
            cm += 1
            if cm > 12:
                cm = 1
                cy += 1
        return periods

    raise ValueError(f"Unknown timeframe: {timeframe}")


def timeframe_label(
    timeframe: TimeframeType,
    periods: list[str],
    quarter: Optional[int] = None,
    year: Optional[int] = None,
) -> str:
    """Human-readable label for the timeframe."""
    if timeframe == TimeframeType.single_month:
        return periods[0]
    if timeframe == TimeframeType.quarter:
        return f"Q{quarter} {year}"
    if timeframe == TimeframeType.ytd:
        return f"Year to date {year or date.today().year}"
    if timeframe == TimeframeType.full_year:
        return f"Full year {year}"
    if timeframe == TimeframeType.custom:
        return f"{periods[0]} – {periods[-1]}"
    return "Report"


def _fmt(value) -> str:
    """Format a Decimal/float for display."""
    if isinstance(value, Decimal):
        return f"{value:,.2f}"
    if isinstance(value, (int, float)):
        return f"{value:,.2f}"
    return str(value)


def _invoice_filename(inv) -> str:
    """Extract original filename from raw_extraction, fallback to invoice_number."""
    if inv.raw_extraction and isinstance(inv.raw_extraction, dict):
        fname = inv.raw_extraction.get("filename") or inv.raw_extraction.get("file_name")
        if fname:
            return fname
    return inv.invoice_number


# ---------------------------------------------------------------------------
# Data aggregation (transaction-centric)
# ---------------------------------------------------------------------------


async def _gather_report_data(db: AsyncSession, periods: list[str]) -> dict:
    """Query all transactions in periods and classify by status."""
    # 1. Load all transactions in the requested periods
    tx_result = await db.execute(
        select(Transaction)
        .where(Transaction.period.in_(periods))
        .order_by(Transaction.tx_date)
    )
    all_txs = tx_result.scalars().all()

    # 2. Classify by debit/credit and status
    debit_txs = [tx for tx in all_txs if tx.amount < 0]
    credit_txs = [tx for tx in all_txs if tx.amount > 0]

    matched_txs = [tx for tx in debit_txs if tx.status == TransactionStatus.matched]
    no_invoice_txs = [tx for tx in debit_txs if tx.status == TransactionStatus.no_invoice]
    withholding_txs = [tx for tx in debit_txs if tx.status == TransactionStatus.withholding]
    unmatched_txs = [tx for tx in debit_txs if tx.status == TransactionStatus.unmatched]

    # 3. Load matches for matched transactions (eager-load invoice + transaction)
    matches: list[Match] = []
    if matched_txs:
        matched_ids = [tx.id for tx in matched_txs]
        match_result = await db.execute(
            select(Match)
            .where(Match.transaction_id.in_(matched_ids))
            .options(selectinload(Match.invoice), selectinload(Match.transaction))
            .order_by(Match.transaction_id)
        )
        matches = list(match_result.scalars().all())
        # Sort matches by transaction date for display
        matches.sort(key=lambda m: m.transaction.tx_date if m.transaction else date.min)

    # 4. Compute summary
    total_debits = sum(tx.amount for tx in debit_txs)
    total_credits = sum(tx.amount for tx in credit_txs)
    matched_total = sum(tx.amount for tx in matched_txs)
    no_invoice_total = sum(tx.amount for tx in no_invoice_txs)
    withholding_total = sum(tx.amount for tx in withholding_txs)
    unmatched_total = sum(tx.amount for tx in unmatched_txs)

    summary = {
        "total_debits": total_debits,
        "total_credits": total_credits,
        "debit_count": len(debit_txs),
        "credit_count": len(credit_txs),
        "matched_count": len(matched_txs),
        "matched_total": matched_total,
        "no_invoice_count": len(no_invoice_txs),
        "no_invoice_total": no_invoice_total,
        "withholding_count": len(withholding_txs),
        "withholding_total": withholding_total,
        "unmatched_count": len(unmatched_txs),
        "unmatched_total": unmatched_total,
    }

    return {
        "summary": summary,
        "matches": matches,
        "no_invoice_txs": no_invoice_txs,
        "withholding_txs": withholding_txs,
        "earnings_txs": credit_txs,
        "unmatched_txs": unmatched_txs,
    }


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------


def _render_html(label: str, data: dict, slim: bool = False) -> str:
    """Build the HTML string for the PDF report."""
    today = date.today().strftime("%Y-%m-%d")
    s = data["summary"]

    report_title = "OpenReconFi — Summary Report" if slim else "OpenReconFi — Financial Report"

    html_parts = [
        f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: 'Helvetica Neue', Arial, sans-serif; margin: 40px; color: #333; font-size: 11px; }}
  h1 {{ color: #1a1a2e; border-bottom: 2px solid #1a1a2e; padding-bottom: 8px; }}
  h2 {{ color: #16213e; margin-top: 30px; border-bottom: 1px solid #ccc; padding-bottom: 4px; }}
  table {{ border-collapse: collapse; width: 100%; margin: 10px 0 20px; }}
  th, td {{ border: 1px solid #ddd; padding: 6px 10px; text-align: left; }}
  th {{ background: #f4f4f8; font-weight: 600; }}
  td.num {{ text-align: right; }}
  .cover {{ text-align: center; margin-top: 120px; }}
  .cover h1 {{ font-size: 28px; border: none; }}
  .cover p {{ font-size: 14px; color: #666; }}
  .page-break {{ page-break-before: always; }}
  .landscape {{ page-break-before: always; page: landscape; }}
  @page landscape {{ size: A4 landscape; margin: 20mm; }}
</style>
</head><body>

<div class="cover">
  <h1>{report_title}</h1>
  <p><strong>{label}</strong></p>
  <p>Generated {today}</p>
</div>

<div class="page-break"></div>
""",
    ]

    if not slim:
        html_parts.append(f"""
<h2>Summary</h2>
<table>
  <tr><th>Metric</th><th>Value</th></tr>
  <tr><td>Total debits</td><td class="num">&euro;{_fmt(s['total_debits'])}</td></tr>
  <tr><td>Total credits</td><td class="num">&euro;{_fmt(s['total_credits'])}</td></tr>
  <tr><td>Debit transactions</td><td class="num">{s['debit_count']}</td></tr>
  <tr><td>Credit transactions</td><td class="num">{s['credit_count']}</td></tr>
  <tr><td>Matched expenses</td><td class="num">{s['matched_count']} (&euro;{_fmt(s['matched_total'])})</td></tr>
  <tr><td>Dismissed (no invoice)</td><td class="num">{s['no_invoice_count']} (&euro;{_fmt(s['no_invoice_total'])})</td></tr>
  <tr><td>Withholdings</td><td class="num">{s['withholding_count']} (&euro;{_fmt(s['withholding_total'])})</td></tr>
  <tr><td>Unmatched expenses</td><td class="num">{s['unmatched_count']} (&euro;{_fmt(s['unmatched_total'])})</td></tr>
</table>""")

    # Matched Expenses
    if data["matches"]:
        if slim:
            html_parts.append("""
<div class="landscape">
<h2>Matched Expenses</h2>
<table>
  <tr>
    <th>Tx Date</th><th>Inv Date</th><th>Amount</th><th>FX Amount</th>
    <th>Vendor</th><th>Invoice #</th><th>Inv Amount</th><th>Currency</th>
    <th>Category</th>
  </tr>""")
            for m in data["matches"]:
                tx = m.transaction
                inv = m.invoice
                fx_amount = f"{_fmt(tx.original_amount)} {tx.original_currency}" if tx.original_amount else ""
                html_parts.append(
                    f"  <tr><td>{tx.tx_date}</td>"
                    f"<td>{inv.invoice_date}</td>"
                    f'<td class="num">&euro;{_fmt(tx.amount)}</td>'
                    f'<td class="num">{fx_amount}</td>'
                    f"<td>{inv.vendor}</td>"
                    f"<td>{inv.invoice_number}</td>"
                    f'<td class="num">{_fmt(inv.amount_incl)}</td>'
                    f"<td>{inv.currency}</td>"
                    f"<td>{inv.category or ''}</td></tr>"
                )
            html_parts.append("</table></div>")
        else:
            html_parts.append("""
<div class="landscape">
<h2>Matched Expenses</h2>
<table>
  <tr>
    <th>Tx Date</th><th>Inv Date</th><th>EUR Amount</th><th>FX Amount</th><th>Counterparty</th>
    <th>Vendor</th><th>Invoice #</th><th>Inv Amount</th><th>Currency</th>
    <th>File</th><th>Category</th><th>Confidence</th><th>Confirmed By</th>
  </tr>""")
            for m in data["matches"]:
                tx = m.transaction
                inv = m.invoice
                fx_amount = f"{_fmt(tx.original_amount)} {tx.original_currency}" if tx.original_amount else ""
                html_parts.append(
                    f"  <tr><td>{tx.tx_date}</td>"
                    f"<td>{inv.invoice_date}</td>"
                    f'<td class="num">&euro;{_fmt(tx.amount)}</td>'
                    f'<td class="num">{fx_amount}</td>'
                    f"<td>{tx.counterparty}</td>"
                    f"<td>{inv.vendor}</td>"
                    f"<td>{inv.invoice_number}</td>"
                    f'<td class="num">{_fmt(inv.amount_incl)}</td>'
                    f"<td>{inv.currency}</td>"
                    f"<td>{_invoice_filename(inv)}</td>"
                    f"<td>{inv.category or ''}</td>"
                    f'<td class="num">{_fmt(m.confidence)}</td>'
                    f"<td>{m.confirmed_by.value}</td></tr>"
                )
            html_parts.append("</table></div>")

    # Withholdings
    if data["withholding_txs"]:
        html_parts.append("""
<h2>Withholdings</h2>
<table>
  <tr><th>Date</th><th>Amount</th><th>Counterparty</th><th>Description</th></tr>""")
        for tx in data["withholding_txs"]:
            html_parts.append(
                f"  <tr><td>{tx.tx_date}</td>"
                f'<td class="num">&euro;{_fmt(tx.amount)}</td>'
                f"<td>{tx.counterparty}</td>"
                f"<td>{tx.description}</td></tr>"
            )
        html_parts.append("</table>")

    # Dismissed (no_invoice)
    if data["no_invoice_txs"]:
        if slim:
            html_parts.append("""
<h2>Dismissed Expenses</h2>
<table>
  <tr><th>Date</th><th>Amount</th><th>Counterparty</th><th>Category</th></tr>""")
            for tx in data["no_invoice_txs"]:
                html_parts.append(
                    f"  <tr><td>{tx.tx_date}</td>"
                    f'<td class="num">&euro;{_fmt(tx.amount)}</td>'
                    f"<td>{tx.counterparty}</td>"
                    f"<td>{tx.category or ''}</td></tr>"
                )
            html_parts.append("</table>")
        else:
            html_parts.append("""
<div class="landscape">
<h2>Dismissed (No Invoice Required)</h2>
<table>
  <tr><th>Date</th><th>Amount</th><th>Counterparty</th><th>Description</th><th>Category</th><th>Note</th></tr>""")
            for tx in data["no_invoice_txs"]:
                html_parts.append(
                    f"  <tr><td>{tx.tx_date}</td>"
                    f'<td class="num">&euro;{_fmt(tx.amount)}</td>'
                    f"<td>{tx.counterparty}</td>"
                    f"<td>{tx.description}</td>"
                    f"<td>{tx.category or ''}</td>"
                    f"<td>{tx.note or ''}</td></tr>"
                )
            html_parts.append("</table></div>")

    # Earnings (credits)
    if data["earnings_txs"]:
        html_parts.append("""
<h2>Earnings</h2>
<table>
  <tr><th>Date</th><th>Amount</th><th>Counterparty</th><th>Description</th></tr>""")
        for tx in data["earnings_txs"]:
            html_parts.append(
                f"  <tr><td>{tx.tx_date}</td>"
                f'<td class="num">&euro;{_fmt(tx.amount)}</td>'
                f"<td>{tx.counterparty}</td>"
                f"<td>{tx.description}</td></tr>"
            )
        html_parts.append("</table>")

    # Unmatched Expenses (full report only)
    if not slim and data["unmatched_txs"]:
        html_parts.append("""
<h2>Unmatched Expenses</h2>
<table>
  <tr><th>Date</th><th>Amount</th><th>Counterparty</th><th>Description</th></tr>""")
        for tx in data["unmatched_txs"]:
            html_parts.append(
                f"  <tr><td>{tx.tx_date}</td>"
                f'<td class="num">&euro;{_fmt(tx.amount)}</td>'
                f"<td>{tx.counterparty}</td>"
                f"<td>{tx.description}</td></tr>"
            )
        html_parts.append("</table>")

    html_parts.append("</body></html>")
    return "\n".join(html_parts)


async def generate_pdf(db: AsyncSession, label: str, periods: list[str], slim: bool = False) -> bytes:
    """Generate a PDF report for the given periods."""
    data = await _gather_report_data(db, periods)
    html_str = _render_html(label, data, slim=slim)
    from weasyprint import HTML  # lazy import — requires system libs (pango, cairo)

    pdf_bytes = HTML(string=html_str).write_pdf()
    return pdf_bytes

def _write_header(ws, headers: list[str]) -> None:
    """Write a styled header row."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


async def generate_excel(db: AsyncSession, label: str, periods: list[str], slim: bool = False) -> bytes:
    """Generate an Excel report for the given periods."""
    data = await _gather_report_data(db, periods)
    wb = openpyxl.Workbook()
    s = data["summary"]

    # --- Summary sheet (full only) ---
    ws = wb.active
    if slim:
        ws.title = "Matched Expenses"
        if data["matches"]:
            _write_header(ws, [
                "Tx Date", "Inv Date", "Amount", "FX Amount", "FX Currency",
                "Vendor", "Invoice #", "Inv Amount", "Currency", "Category",
            ])
            for i, m in enumerate(data["matches"], 2):
                tx = m.transaction
                inv = m.invoice
                ws.cell(row=i, column=1, value=str(tx.tx_date))
                ws.cell(row=i, column=2, value=str(inv.invoice_date))
                ws.cell(row=i, column=3, value=float(tx.amount))
                ws.cell(row=i, column=4, value=float(tx.original_amount) if tx.original_amount else None)
                ws.cell(row=i, column=5, value=tx.original_currency or "")
                ws.cell(row=i, column=6, value=inv.vendor)
                ws.cell(row=i, column=7, value=inv.invoice_number)
                ws.cell(row=i, column=8, value=float(inv.amount_incl))
                ws.cell(row=i, column=9, value=inv.currency)
                ws.cell(row=i, column=10, value=inv.category or "")
    else:
        ws.title = "Summary"
        _write_header(ws, ["Metric", "Value"])
        rows = [
            ("Total debits", float(s["total_debits"])),
            ("Total credits", float(s["total_credits"])),
            ("Debit transactions", s["debit_count"]),
            ("Credit transactions", s["credit_count"]),
            ("Matched expenses", s["matched_count"]),
            ("Matched total", float(s["matched_total"])),
            ("Dismissed (no invoice)", s["no_invoice_count"]),
            ("Dismissed total", float(s["no_invoice_total"])),
            ("Withholdings", s["withholding_count"]),
            ("Withholding total", float(s["withholding_total"])),
            ("Unmatched expenses", s["unmatched_count"]),
            ("Unmatched total", float(s["unmatched_total"])),
        ]
        for i, (metric, val) in enumerate(rows, 2):
            ws.cell(row=i, column=1, value=metric)
            ws.cell(row=i, column=2, value=val)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18

    # --- Matched Expenses sheet (full only) ---
    if not slim and data["matches"]:
        ws_m = wb.create_sheet("Matched Expenses")
        _write_header(ws_m, [
            "Tx Date", "Inv Date", "EUR Amount", "FX Amount", "FX Currency", "Counterparty",
            "Vendor", "Invoice #", "Inv Amount", "Inv Currency",
            "File", "Category", "Confidence", "Confirmed By",
        ])
        for i, m in enumerate(data["matches"], 2):
            tx = m.transaction
            inv = m.invoice
            ws_m.cell(row=i, column=1, value=str(tx.tx_date))
            ws_m.cell(row=i, column=2, value=str(inv.invoice_date))
            ws_m.cell(row=i, column=3, value=float(tx.amount))
            ws_m.cell(row=i, column=4, value=float(tx.original_amount) if tx.original_amount else None)
            ws_m.cell(row=i, column=5, value=tx.original_currency or "")
            ws_m.cell(row=i, column=6, value=tx.counterparty)
            ws_m.cell(row=i, column=7, value=inv.vendor)
            ws_m.cell(row=i, column=8, value=inv.invoice_number)
            ws_m.cell(row=i, column=9, value=float(inv.amount_incl))
            ws_m.cell(row=i, column=10, value=inv.currency)
            ws_m.cell(row=i, column=11, value=_invoice_filename(inv))
            ws_m.cell(row=i, column=12, value=inv.category or "")
            ws_m.cell(row=i, column=13, value=float(m.confidence))
            ws_m.cell(row=i, column=14, value=m.confirmed_by.value)

    # --- Withholdings sheet ---
    if data["withholding_txs"]:
        ws_w = wb.create_sheet("Withholdings")
        _write_header(ws_w, ["Date", "Amount", "Counterparty", "Description"])
        for i, tx in enumerate(data["withholding_txs"], 2):
            ws_w.cell(row=i, column=1, value=str(tx.tx_date))
            ws_w.cell(row=i, column=2, value=float(tx.amount))
            ws_w.cell(row=i, column=3, value=tx.counterparty)
            ws_w.cell(row=i, column=4, value=tx.description)

    # --- Dismissed sheet ---
    if data["no_invoice_txs"]:
        ws_d = wb.create_sheet("Dismissed")
        if slim:
            _write_header(ws_d, ["Date", "Amount", "Counterparty", "Category"])
            for i, tx in enumerate(data["no_invoice_txs"], 2):
                ws_d.cell(row=i, column=1, value=str(tx.tx_date))
                ws_d.cell(row=i, column=2, value=float(tx.amount))
                ws_d.cell(row=i, column=3, value=tx.counterparty)
                ws_d.cell(row=i, column=4, value=tx.category or "")
        else:
            _write_header(ws_d, ["Date", "Amount", "Counterparty", "Description", "Category", "Note"])
            for i, tx in enumerate(data["no_invoice_txs"], 2):
                ws_d.cell(row=i, column=1, value=str(tx.tx_date))
                ws_d.cell(row=i, column=2, value=float(tx.amount))
                ws_d.cell(row=i, column=3, value=tx.counterparty)
                ws_d.cell(row=i, column=4, value=tx.description)
                ws_d.cell(row=i, column=5, value=tx.category or "")
                ws_d.cell(row=i, column=6, value=tx.note or "")

    # --- Earnings sheet ---
    if data["earnings_txs"]:
        ws_e = wb.create_sheet("Earnings")
        _write_header(ws_e, ["Date", "Amount", "Counterparty", "Description"])
        for i, tx in enumerate(data["earnings_txs"], 2):
            ws_e.cell(row=i, column=1, value=str(tx.tx_date))
            ws_e.cell(row=i, column=2, value=float(tx.amount))
            ws_e.cell(row=i, column=3, value=tx.counterparty)
            ws_e.cell(row=i, column=4, value=tx.description)

    # --- Unmatched Expenses sheet (full only) ---
    if not slim and data["unmatched_txs"]:
        ws_u = wb.create_sheet("Unmatched Expenses")
        _write_header(ws_u, ["Date", "Amount", "Counterparty", "Description"])
        for i, tx in enumerate(data["unmatched_txs"], 2):
            ws_u.cell(row=i, column=1, value=str(tx.tx_date))
            ws_u.cell(row=i, column=2, value=float(tx.amount))
            ws_u.cell(row=i, column=3, value=tx.counterparty)
            ws_u.cell(row=i, column=4, value=tx.description)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
